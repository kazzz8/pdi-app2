#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
現場作業管理システム（仮称）機能一覧表
Excel生成スクリプト
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "feature_list.xlsx")

# ============================================================
# カラーパレット
# ============================================================
C_NAVY     = "1A3A6B"
C_TEAL     = "0E6E75"
C_GREEN    = "1E8B4C"
C_ORANGE   = "C06000"
C_RED      = "C0392B"
C_PURPLE   = "5B2D8E"
C_GRAY_H   = "4A4A4A"
C_GRAY_BG  = "F2F4F8"
C_LIGHT    = "EAF0FB"
C_WHITE    = "FFFFFF"
C_BORDER   = "CCCCCC"
C_DONE_BG  = "E8F5E9"   # 実装済み行背景
C_WIP_BG   = "FFF8E1"   # 実装中行背景
C_PLAN_BG  = "F5F5F5"   # 計画中行背景

# 大分類ごとの色
CATEGORY_COLOR = {
    "認証・セッション":       C_NAVY,
    "ダッシュボード":          C_TEAL,
    "作業開始・車両確認":      C_GREEN,
    "作業中・タイマー管理":    C_GREEN,
    "点検・不具合記録":        C_RED,
    "作業完了・報告書":        C_TEAL,
    "管理者機能":              C_PURPLE,
    "班長機能":                C_ORANGE,
}

# 実装状況
DONE  = "✅ 実装済み"
WIP   = "🔧 実装中"
PLAN  = "📋 計画中"

# ロール: ○=利用可, —=利用不可
O = "○"
D = "—"

# ============================================================
# 機能定義データ
# 列: (大分類, 機能名, 機能説明, 画面URL, メソッド, APIエンドポイント,
#      GENERAL, TEAM_LEADER, MANAGER, 実装状況, 備考)
# ============================================================
FEATURES = [

    # ---- 認証・セッション ----
    ("認証・セッション", "ログイン",
     "社員番号とパスワードでシステムにログインする",
     "/login", "POST", "/api/auth/callback/credentials",
     O, O, O, DONE, "bcryptでパスワードを照合。NextAuth.js使用"),

    ("認証・セッション", "ログアウト",
     "セッションを破棄してログイン画面に戻る",
     "/dashboard", "POST", "/api/auth/signout",
     O, O, O, DONE, "ダッシュボードのヘッダーボタンから操作"),

    ("認証・セッション", "セッション維持（JWT）",
     "ログイン状態をJWTトークンで維持する",
     "全画面", "—", "—",
     O, O, O, DONE, "社員ID・氏名・ロール・班情報をトークンに格納"),

    ("認証・セッション", "未認証リダイレクト",
     "未ログイン状態でのアクセスをログイン画面に転送する",
     "全画面", "—", "—",
     O, O, O, DONE, "middleware.tsで全ルートを保護"),

    # ---- ダッシュボード ----
    ("ダッシュボード", "本日の作業計画一覧表示",
     "自分に割り当てられた当日分の作業計画を一覧で表示する",
     "/dashboard", "GET", "/api/dashboard/my-plans",
     O, O, O, DONE, "計画開始時刻の昇順で表示"),

    ("ダッシュボード", "作業ステータス表示",
     "各作業計画の進行状況をバッジで表示する",
     "/dashboard", "—", "—",
     O, O, O, DONE, "完了 / 作業中 / 中断中 / 遅延 / 予定通り の5種類"),

    ("ダッシュボード", "遅延自動判定",
     "計画終了時刻を過ぎた未完了作業を「遅延」として自動判定する",
     "/dashboard", "—", "—",
     O, O, O, DONE, "サーバー側で現在時刻と計画終了時刻を比較"),

    ("ダッシュボード", "中断中バナー表示",
     "中断中の作業がある場合、画面上部にバナーを表示する",
     "/dashboard", "—", "—",
     O, O, O, DONE, "activeInterruptionLogIdが存在する場合に表示"),

    ("ダッシュボード", "作業開始ナビゲーション",
     "作業計画の「開始」ボタンを押して作業開始確認画面に遷移する",
     "/dashboard", "—", "—",
     O, O, O, DONE, ""),

    ("ダッシュボード", "中断作業の再開",
     "「再開」ボタンから中断中の作業を継続する",
     "/dashboard", "POST", "/api/work/[planId]/resume",
     O, O, O, DONE, "中断ログIDをPOSTして作業中画面へ遷移"),

    ("ダッシュボード", "計画外作業の開始",
     "作業計画にない突発作業を手動で開始する",
     "/dashboard", "POST", "（未実装）",
     O, O, O, WIP,  "UIボタンは実装済み。API・遷移先は未実装"),

    # ---- 作業開始・車両確認 ----
    ("作業開始・車両確認", "作業詳細表示",
     "対象作業の工程・計画時刻・車両情報（車種・色・整理番号）を表示する",
     "/work/[planId]", "GET", "/api/work/[planId]",
     O, O, O, DONE, ""),

    ("作業開始・車両確認", "バーコードスキャン起動",
     "スマートフォンのカメラを起動してバーコードを読み取る",
     "/work/[planId]", "—", "—",
     O, O, O, DONE, "BarcodeScanner コンポーネントを使用"),

    ("作業開始・車両確認", "車両バーコード照合",
     "スキャンしたバーコードと作業計画の整理番号が一致するか確認する",
     "/work/[planId]", "—", "—",
     O, O, O, DONE, "一致時のみ「作業開始」ボタンが活性化"),

    ("作業開始・車両確認", "バーコード不一致エラー表示",
     "スキャン結果が計画車両と一致しない場合にエラーを表示する",
     "/work/[planId]", "—", "—",
     O, O, O, DONE, "再スキャンボタンを表示"),

    ("作業開始・車両確認", "作業開始（WorkLog作成）",
     "バーコード確認後に作業を開始し、作業実績レコードを作成する",
     "/work/[planId]", "POST", "/api/work/[planId]/start",
     O, O, O, DONE, "WorkLogをステータスACTIVEで作成し作業中画面へ遷移"),

    # ---- 作業中・タイマー管理 ----
    ("作業中・タイマー管理", "リアルタイムタイマー表示",
     "作業の経過時間をMM:SS形式でリアルタイム表示する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "1秒ごとに更新。累積時間+今回の経過時間を合算"),

    ("作業中・タイマー管理", "累積作業時間取得",
     "中断・再開を繰り返した場合の通算作業時間をサーバーから取得する",
     "/work/[planId]/active", "GET", "/api/work/[planId]/elapsed",
     O, O, O, DONE, "過去のWORKLOG完了レコードから累積秒数を算出"),

    ("作業中・タイマー管理", "標準時間超過警告",
     "標準作業時間を超過した場合にタイマーを赤色にして警告を表示する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "現在は20分固定。WorkPlanのstandardMinutesとの連携は今後"),

    ("作業中・タイマー管理", "一時中断",
     "作業を一時停止してダッシュボードに戻る",
     "/work/[planId]/active", "POST", "/api/work/[planId]/pause",
     O, O, O, DONE, "WorkLogをPAUSEDに更新。中断理由も記録"),

    ("作業中・タイマー管理", "中断理由クイック選択",
     "定型の中断理由（班長対応・緊急補修・部品待ち・休憩・その他）をボタン選択する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, ""),

    ("作業中・タイマー管理", "中断理由自由入力",
     "定型以外の中断理由をテキストで入力する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, ""),

    ("作業中・タイマー管理", "作業再開",
     "中断中の作業を再開し、新しい作業実績レコードを作成する",
     "/dashboard", "POST", "/api/work/[planId]/resume",
     O, O, O, DONE, "中断ログのIDをパラメータとして送信"),

    # ---- 点検・不具合記録 ----
    ("点検・不具合記録", "車両展開図表示",
     "L点検・S点検作業中に車両の展開図（全面）をインタラクティブに表示する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "VehicleDiagramコンポーネント。上面/前後/左右側面を1画面に表示"),

    ("点検・不具合記録", "展開図タップで不具合位置指定",
     "展開図上の任意の位置をタップして不具合の発生箇所を指定する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "タップ座標をX%・Y%で記録"),

    ("点検・不具合記録", "不具合種類選択",
     "不具合の種類（線傷/面傷/塗装/へこみ/欠け/汚れ/その他）をボタンで選択する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "必須項目"),

    ("点検・不具合記録", "不具合程度選択",
     "不具合の程度をA（軽微）/B（中程度）/C（重大）から選択する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "選択結果が展開図ピンの色に反映される"),

    ("点検・不具合記録", "補修見込み時間入力",
     "補修に要する見込み時間をプリセット（10/30/60/90/120分）または5分刻みで入力する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "+/−ボタンによる微調整も可能"),

    ("点検・不具合記録", "不具合写真撮影",
     "スマートフォンのリアカメラを起動して不具合箇所を撮影する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "任意項目。capture=\"environment\"でリアカメラを指定"),

    ("点検・不具合記録", "写真リサイズ・圧縮",
     "撮影画像を最大800pxにリサイズしてJPEG 70%品質に圧縮する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "Canvas APIでブラウザ内処理。通信量・容量を削減"),

    ("点検・不具合記録", "不具合データのローカル保存",
     "入力中の不具合情報をlocalStorageに自動保存する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "ブラウザを閉じた場合や中断時のデータ消失を防止"),

    ("点検・不具合記録", "不具合ピンの展開図表示",
     "登録した不具合を展開図上にピン（番号付き色付き丸）で表示する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "程度A=黄, B=橙, C=赤で色分け"),

    ("点検・不具合記録", "不具合一覧表示",
     "登録した不具合を種類・程度・補修時間・写真とともに一覧表示する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, ""),

    ("点検・不具合記録", "不具合詳細表示",
     "不具合ピンまたは一覧アイテムをタップして詳細情報を表示する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "モーダルで表示"),

    ("点検・不具合記録", "不具合削除",
     "登録した不具合を削除する",
     "/work/[planId]/active", "—", "—",
     O, O, O, DONE, "詳細モーダルから操作"),

    # ---- 作業完了・報告書 ----
    ("作業完了・報告書", "写真一括アップロード",
     "作業完了時にlocalStorageの写真データをサーバーにアップロードする",
     "/work/[planId]/active", "POST", "/api/upload",
     O, O, O, DONE, "FormDataでBlobを送信。サーバー上にファイル保存"),

    ("作業完了・報告書", "作業完了処理",
     "作業実績を完了状態に更新し、点検報告書と不具合データをDBに保存する",
     "/work/[planId]/active", "POST", "/api/work/[planId]/complete",
     O, O, O, DONE, "WorkLog完了 + InspectionReport作成 + Defect一括登録"),

    ("作業完了・報告書", "点検報告書の自動生成",
     "作業完了時にInspectionReportレコードを自動作成する",
     "—（サーバー処理）", "—", "—",
     O, O, O, DONE, "点検種別・開始〜終了時刻・作業者を記録"),

    ("作業完了・報告書", "不具合のDB一括登録",
     "完了時に入力した全不具合をDefectテーブルに一括保存する",
     "—（サーバー処理）", "—", "—",
     O, O, O, DONE, "座標・種類・程度・写真URLを含む"),

    ("作業完了・報告書", "点検完了画面表示",
     "作業完了後に点検報告書のサマリーを確認画面として表示する",
     "/work/[planId]/complete", "GET", "/api/inspection-reports/[reportId]",
     O, O, O, DONE, ""),

    ("作業完了・報告書", "点検報告書サマリー表示",
     "完了画面で車両情報・作業時間・発見した不具合一覧を表示する",
     "/work/[planId]/complete", "—", "—",
     O, O, O, DONE, "不具合0件の場合は「異常なし」と表示"),

    ("作業完了・報告書", "ダッシュボードへ戻る",
     "完了確認後にダッシュボードに戻る",
     "/work/[planId]/complete", "—", "—",
     O, O, O, DONE, ""),

    # ---- 管理者機能（計画中） ----
    ("管理者機能", "ユーザー登録・管理",
     "作業者アカウントの作成・編集・ロール変更・無効化を行う",
     "（未実装）", "—", "—",
     D, D, O, PLAN, "管理者のみ操作可能"),

    ("管理者機能", "班管理",
     "班の作成・編集・メンバー割り当てを管理する",
     "（未実装）", "—", "—",
     D, D, O, PLAN, ""),

    ("管理者機能", "作業計画CSVインポート",
     "作業計画データをCSVファイルから一括インポートする",
     "（未実装）", "POST", "（未実装）",
     D, D, O, PLAN, "ImportBatchテーブルにインポート履歴を記録"),

    ("管理者機能", "車両データ登録",
     "入庫した車両情報をシステムに登録する",
     "（未実装）", "—", "—",
     D, D, O, PLAN, "バーコード・車種・色・点検区分等を登録"),

    ("管理者機能", "保管場所管理",
     "保管ゾーン・ポジションの登録・車両割り当てを管理する",
     "（未実装）", "—", "—",
     D, D, O, PLAN, ""),

    ("管理者機能", "ドレスアップ品マスタ管理",
     "ドレスアップ部品のマスタ情報を登録・編集する",
     "（未実装）", "—", "—",
     D, D, O, PLAN, ""),

    ("管理者機能", "全体進捗ダッシュボード",
     "全チームの作業進捗をリアルタイムで一覧表示する",
     "（未実装）", "—", "—",
     D, O, O, PLAN, "班長・管理者が閲覧可能"),

    ("管理者機能", "作業実績レポート出力",
     "期間・班・工程を指定して作業実績を集計・出力する",
     "（未実装）", "—", "—",
     D, O, O, PLAN, "Excel出力等を想定"),

    # ---- 班長機能（計画中） ----
    ("班長機能", "チームメンバーの進捗確認",
     "自班のメンバーの作業状況をリアルタイムで確認する",
     "（未実装）", "—", "—",
     D, O, O, PLAN, ""),

    ("班長機能", "不具合レポート閲覧",
     "担当車両の不具合発見状況・補修進捗を一覧で確認する",
     "（未実装）", "—", "—",
     D, O, O, PLAN, ""),

    ("班長機能", "作業指示・担当変更",
     "メンバーへの作業割り当てを変更・調整する",
     "（未実装）", "—", "—",
     D, O, O, PLAN, ""),
]

# ============================================================
# スタイルヘルパー
# ============================================================
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(size=10, bold=False, color="000000", name="游ゴシック", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, row, col, value="", bg=None, fnt=None, aln=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:  c.fill = fill(bg)
    if fnt: c.font = fnt
    if aln: c.alignment = aln
    c.border = border()
    return c


# ============================================================
# 機能一覧シート
# ============================================================
def make_feature_sheet(wb):
    ws = wb.create_sheet("機能一覧")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # 列定義: (幅, ヘッダー)
    columns = [
        (5,  "No."),
        (20, "大分類"),
        (24, "機能名"),
        (40, "機能説明"),
        (28, "対応画面（URL）"),
        (8,  "HTTP"),
        (35, "APIエンドポイント"),
        (8,  "一般"),
        (8,  "班長"),
        (8,  "管理者"),
        (14, "実装状況"),
        (30, "備考"),
    ]

    for i, (w, _) in enumerate(columns):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # タイトル行
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    c = ws.cell(1, 1, "現場作業管理システム（仮称）　機能一覧表")
    c.fill = fill(C_NAVY)
    c.font = font(size=14, bold=True, color=C_WHITE)
    c.alignment = align("left", "center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(len(columns))}2")
    c2 = ws.cell(2, 1, "2026年4月現在　　✅ 実装済み　　🔧 実装中　　📋 計画中")
    c2.fill = fill(C_LIGHT)
    c2.font = font(size=10, color=C_NAVY)
    c2.alignment = align("right", "center")
    ws.row_dimensions[2].height = 18

    # ヘッダー行
    for i, (_, h) in enumerate(columns):
        cell_style(ws, 3, i + 1, h,
                   bg=C_NAVY,
                   fnt=font(bold=True, color=C_WHITE, size=10),
                   aln=align("center", "center"))
    ws.row_dimensions[3].height = 22

    # データ行
    prev_cat = None
    no = 0

    for row_i, feat in enumerate(FEATURES):
        (cat, name, desc, screen, method, endpoint,
         gen, tl, mgr, status, notes) = feat

        r = 4 + row_i
        no += 1
        cat_color = CATEGORY_COLOR.get(cat, C_GRAY_H)

        # 実装状況で背景色を変える
        if status == DONE:
            base_bg = C_DONE_BG if row_i % 2 == 0 else C_WHITE
        elif status == WIP:
            base_bg = C_WIP_BG
        else:
            base_bg = C_PLAN_BG

        # 大分類が変わったら色帯を入れる
        is_new_cat = (cat != prev_cat)
        row_bg = base_bg

        cell_style(ws, r, 1, no,
                   bg=row_bg, fnt=font(size=9, color="888888"),
                   aln=align("center", "center"))

        # 大分類セル
        cell_style(ws, r, 2, cat if is_new_cat else "",
                   bg=cat_color if is_new_cat else row_bg,
                   fnt=font(bold=True, color=C_WHITE if is_new_cat else cat_color, size=10),
                   aln=align("center", "center"))

        cell_style(ws, r, 3, name,
                   bg=row_bg, fnt=font(bold=True, color=cat_color),
                   aln=align("left", "center"))

        cell_style(ws, r, 4, desc,
                   bg=row_bg, fnt=font(size=10),
                   aln=align("left", "center", wrap=True))

        cell_style(ws, r, 5, screen,
                   bg=row_bg, fnt=font(size=9, color="555555", italic=(screen == "（未実装）")),
                   aln=align("left", "center"))

        cell_style(ws, r, 6, method,
                   bg=row_bg, fnt=font(size=9, color="1A3A6B" if method not in ("—", "") else "AAAAAA", bold=(method not in ("—", ""))),
                   aln=align("center", "center"))

        cell_style(ws, r, 7, endpoint,
                   bg=row_bg, fnt=font(size=9, color="0E6E75" if endpoint not in ("—", "（未実装）", "") else "AAAAAA"),
                   aln=align("left", "center"))

        # ロール列
        for ci, (val, col_i) in enumerate([(gen, 8), (tl, 9), (mgr, 10)]):
            cell_style(ws, r, col_i, val,
                       bg=row_bg,
                       fnt=font(bold=(val == O), color=C_GREEN if val == O else "BBBBBB", size=11),
                       aln=align("center", "center"))

        # 実装状況
        status_bg = {"✅ 実装済み": "D4EDDA", "🔧 実装中": "FFF3CD", "📋 計画中": "E2E3E5"}.get(status, row_bg)
        status_fc = {"✅ 実装済み": "155724", "🔧 実装中": "856404", "📋 計画中": "383D41"}.get(status, "000000")
        cell_style(ws, r, 11, status,
                   bg=status_bg,
                   fnt=font(bold=True, color=status_fc, size=9),
                   aln=align("center", "center"))

        cell_style(ws, r, 12, notes,
                   bg=row_bg, fnt=font(size=9, color="555555", italic=True),
                   aln=align("left", "center", wrap=True))

        ws.row_dimensions[r].height = 30
        prev_cat = cat


# ============================================================
# 画面・API対応表シート
# ============================================================
def make_api_sheet(wb):
    ws = wb.create_sheet("画面・API一覧")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    screens = [
        ("/login",                    "ログイン画面",         ["POST /api/auth/callback/credentials"],         "認証・セッション", DONE),
        ("/dashboard",                "ダッシュボード",        ["GET /api/dashboard/my-plans",
                                                               "POST /api/work/[planId]/resume"],              "ダッシュボード",    DONE),
        ("/work/[planId]",            "作業開始確認画面",      ["GET /api/work/[planId]",
                                                               "POST /api/work/[planId]/start"],               "作業開始・車両確認",DONE),
        ("/work/[planId]/active",     "作業中画面",            ["GET /api/work/[planId]/elapsed",
                                                               "POST /api/work/[planId]/pause",
                                                               "POST /api/work/[planId]/complete",
                                                               "POST /api/upload"],                            "作業中・不具合記録",DONE),
        ("/work/[planId]/complete",   "作業完了・報告書画面",  ["GET /api/inspection-reports/[reportId]"],      "作業完了・報告書",  DONE),
        ("（未実装）/admin/users",    "ユーザー管理画面",      ["CRUD /api/admin/users"],                       "管理者機能",        PLAN),
        ("（未実装）/admin/plans",    "作業計画管理画面",      ["POST /api/admin/plans/import"],                "管理者機能",        PLAN),
        ("（未実装）/admin/dashboard","管理者ダッシュボード",  ["GET /api/admin/progress"],                     "管理者・班長機能",  PLAN),
    ]

    col_widths = [32, 22, 55, 22, 14]
    headers    = ["画面URL", "画面名", "使用するAPIエンドポイント", "機能カテゴリ", "実装状況"]

    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(1, 1, "画面・APIエンドポイント対応表")
    c.fill = fill(C_TEAL)
    c.font = font(size=13, bold=True, color=C_WHITE)
    c.alignment = align("left", "center")
    ws.row_dimensions[1].height = 26

    for i, h in enumerate(headers):
        cell_style(ws, 2, i + 1, h,
                   bg=C_TEAL,
                   fnt=font(bold=True, color=C_WHITE),
                   aln=align("center", "center"))
    ws.row_dimensions[2].height = 20

    for ri, (url, name, apis, cat, status) in enumerate(screens):
        r = 3 + ri
        bg = C_DONE_BG if status == DONE else C_PLAN_BG

        cell_style(ws, r, 1, url,
                   bg=bg, fnt=font(size=10, color=C_TEAL, bold=True),
                   aln=align("left", "center"))
        cell_style(ws, r, 2, name,
                   bg=bg, fnt=font(bold=True),
                   aln=align("left", "center"))
        cell_style(ws, r, 3, "\n".join(apis),
                   bg=bg, fnt=font(size=9, color="0E6E75"),
                   aln=align("left", "center", wrap=True))

        cat_color = CATEGORY_COLOR.get(cat, C_GRAY_H)
        cell_style(ws, r, 4, cat,
                   bg=bg, fnt=font(color=cat_color, bold=True, size=9),
                   aln=align("center", "center"))

        status_bg = "D4EDDA" if status == DONE else "E2E3E5"
        status_fc = "155724" if status == DONE else "383D41"
        cell_style(ws, r, 5, status,
                   bg=status_bg, fnt=font(bold=True, color=status_fc, size=9),
                   aln=align("center", "center"))

        ws.row_dimensions[r].height = max(22, 18 * len(apis))


# ============================================================
# 集計シート
# ============================================================
def make_summary_sheet(wb):
    ws = wb.create_sheet("集計")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 3

    ws.merge_cells("B1:E1")
    c = ws.cell(1, 1)
    ws.merge_cells("A1:F1")
    c2 = ws.cell(1, 1, "機能実装状況サマリー")
    c2.fill = fill(C_NAVY)
    c2.font = font(size=13, bold=True, color=C_WHITE)
    c2.alignment = align("center", "center")
    ws.row_dimensions[1].height = 28

    # 全体集計
    total      = len(FEATURES)
    done_count = sum(1 for f in FEATURES if f[9] == DONE)
    wip_count  = sum(1 for f in FEATURES if f[9] == WIP)
    plan_count = sum(1 for f in FEATURES if f[9] == PLAN)

    summary_data = [
        ("実装済み（✅）",  done_count,  "D4EDDA", "155724"),
        ("実装中（🔧）",    wip_count,   "FFF3CD", "856404"),
        ("計画中（📋）",    plan_count,  "E2E3E5", "383D41"),
        ("合計",            total,       C_LIGHT,  C_NAVY),
    ]

    ws.row_dimensions[2].height = 8

    headers = ["ステータス", "件数", "割合"]
    for i, h in enumerate(headers):
        cell_style(ws, 3, 2 + i, h,
                   bg=C_NAVY, fnt=font(bold=True, color=C_WHITE),
                   aln=align("center", "center"))
    ws.row_dimensions[3].height = 20

    for ri, (label, count, bg, fc) in enumerate(summary_data):
        r = 4 + ri
        cell_style(ws, r, 2, label, bg=bg, fnt=font(bold=True, color=fc), aln=align("left", "center"))
        cell_style(ws, r, 3, count, bg=bg, fnt=font(bold=True, color=fc, size=14), aln=align("center", "center"))
        pct = f"{count / total * 100:.0f}%" if label != "合計" else "100%"
        cell_style(ws, r, 4, pct, bg=bg, fnt=font(color=fc), aln=align("center", "center"))
        ws.row_dimensions[r].height = 24

    # カテゴリ別集計
    ws.row_dimensions[9].height = 12

    cell_style(ws, 10, 2, "カテゴリ別集計",
               bg=C_NAVY, fnt=font(bold=True, color=C_WHITE), aln=align("center", "center"))
    cell_style(ws, 10, 3, "実装済み",
               bg=C_NAVY, fnt=font(bold=True, color=C_WHITE), aln=align("center", "center"))
    cell_style(ws, 10, 4, "実装中",
               bg=C_NAVY, fnt=font(bold=True, color=C_WHITE), aln=align("center", "center"))
    cell_style(ws, 10, 5, "計画中",
               bg=C_NAVY, fnt=font(bold=True, color=C_WHITE), aln=align("center", "center"))
    ws.row_dimensions[10].height = 20

    cats = list(dict.fromkeys(f[0] for f in FEATURES))
    for ci, cat in enumerate(cats):
        r = 11 + ci
        cat_feats = [f for f in FEATURES if f[0] == cat]
        d = sum(1 for f in cat_feats if f[9] == DONE)
        w = sum(1 for f in cat_feats if f[9] == WIP)
        p = sum(1 for f in cat_feats if f[9] == PLAN)
        color = CATEGORY_COLOR.get(cat, C_GRAY_H)

        cell_style(ws, r, 2, cat,
                   bg=color, fnt=font(bold=True, color=C_WHITE), aln=align("center", "center"))
        cell_style(ws, r, 3, d,
                   bg="D4EDDA", fnt=font(color="155724", bold=True), aln=align("center", "center"))
        cell_style(ws, r, 4, w if w else "—",
                   bg="FFF3CD" if w else C_WHITE, fnt=font(color="856404" if w else "AAAAAA"), aln=align("center", "center"))
        cell_style(ws, r, 5, p if p else "—",
                   bg="E2E3E5" if p else C_WHITE, fnt=font(color="383D41" if p else "AAAAAA"), aln=align("center", "center"))
        ws.row_dimensions[r].height = 20


# ============================================================
# メイン
# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)

    make_feature_sheet(wb)
    make_api_sheet(wb)
    make_summary_sheet(wb)

    wb.save(OUT)
    print(f"保存完了: {OUT}")


if __name__ == "__main__":
    main()
