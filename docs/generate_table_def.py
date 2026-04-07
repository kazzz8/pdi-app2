#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
現場作業管理システム（仮称）テーブル定義書
Excel生成スクリプト（prisma/schema.prisma から作成）
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "table_definition.xlsx")

# ============================================================
# カラーパレット
# ============================================================
C_NAVY      = "1A3A6B"
C_NAVY_L    = "2E5FA3"
C_TEAL      = "0E6E75"
C_GREEN     = "1E8B4C"
C_ORANGE    = "C06000"
C_RED       = "C0392B"
C_GRAY_H    = "4A4A4A"  # ヘッダー文字
C_GRAY_BG   = "F2F4F8"  # 偶数行背景
C_LIGHT     = "EAF0FB"  # 薄青
C_YELLOW    = "FFFBE6"  # 注記背景
C_WHITE     = "FFFFFF"
C_BORDER    = "CCCCCC"

# カテゴリカラー
CATEGORY_COLOR = {
    "マスタ系":           C_NAVY,
    "車両・計画系":        C_TEAL,
    "作業実績系":          C_GREEN,
    "点検・補修系":        C_RED,
    "ドレスアップ・チェック系": C_ORANGE,
}

# ============================================================
# テーブル定義データ
# カラム: (物理名, 論理名, 型, NULL可, デフォルト, PK, FK参照先, 備考)
# ============================================================
TABLES = [
    # ---- マスタ系 ----
    {
        "name": "Team", "ja": "班マスタ", "category": "マスタ系",
        "desc": "作業班のマスタ情報。作業者は1つの班に所属する。",
        "columns": [
            ("id",        "ID",       "TEXT",       False, "cuid()", True,  None,        "CUID形式の主キー"),
            ("name",      "班名",     "TEXT",       False, "—",      False, None,        "UNIQUE制約あり"),
            ("createdAt", "作成日時", "TIMESTAMP",  False, "now()",  False, None,        "レコード作成日時"),
        ]
    },
    {
        "name": "User", "ja": "ユーザー", "category": "マスタ系",
        "desc": "システム利用者（作業者・班長・管理者）のアカウント情報。",
        "columns": [
            ("id",         "ID",         "TEXT",      False, "cuid()", True,  None,        "CUID形式の主キー"),
            ("employeeId", "社員番号",   "TEXT",      False, "—",      False, None,        "UNIQUE制約あり。ログインIDとして使用"),
            ("name",       "氏名",       "TEXT",      False, "—",      False, None,        ""),
            ("password",   "パスワード", "TEXT",      False, "—",      False, None,        "bcryptでハッシュ化して保存"),
            ("role",       "ロール",     "ENUM(Role)",False, "GENERAL",False, None,        "GENERAL / TEAM_LEADER / MANAGER"),
            ("teamId",     "班ID",       "TEXT",      True,  "NULL",   False, "Team.id",   "NULL = 未所属"),
            ("createdAt",  "作成日時",   "TIMESTAMP", False, "now()",  False, None,        ""),
        ]
    },
    {
        "name": "DressUpItem", "ja": "ドレスアップ品マスタ", "category": "マスタ系",
        "desc": "ドレスアップ作業で使用する部品・オプションのマスタ。",
        "columns": [
            ("id",            "ID",             "TEXT",  False, "cuid()", True,  None, "CUID形式の主キー"),
            ("name",          "品名",           "TEXT",  False, "—",      False, None, ""),
            ("code",          "品コード",        "TEXT",  False, "—",      False, None, "UNIQUE制約あり"),
            ("checkTemplate", "チェックテンプレート","JSONB", True, "NULL",  False, None, "チェック項目のJSONテンプレート"),
            ("createdAt",     "作成日時",        "TIMESTAMP", False, "now()", False, None, ""),
        ]
    },
    {
        "name": "StorageLocation", "ja": "保管場所マスタ", "category": "マスタ系",
        "desc": "車両の保管場所（ゾーン・ポジション）のマスタ。",
        "columns": [
            ("id",           "ID",           "TEXT",    False, "cuid()", True,  None, "CUID形式の主キー"),
            ("zone",         "ゾーン",       "TEXT",    False, "—",      False, None, ""),
            ("positionCode", "ポジションコード","TEXT",  False, "—",      False, None, ""),
            ("isOccupied",   "使用中フラグ", "BOOLEAN", False, "false",  False, None, "true = 車両あり"),
            ("createdAt",    "作成日時",     "TIMESTAMP",False,"now()",  False, None, ""),
        ]
    },
    # ---- 車両・計画系 ----
    {
        "name": "Vehicle", "ja": "車両", "category": "車両・計画系",
        "desc": "PDI対象の完成車情報。入庫から出庫まで1レコードで管理する。",
        "columns": [
            ("id",                    "ID",               "TEXT",                False, "cuid()", True,  None,                  "CUID形式の主キー"),
            ("serialYear",            "整理番号_年",      "TEXT",                False, "—",      False, None,                  "整理番号の年部分"),
            ("serialMonth",           "整理番号_月",      "TEXT",                False, "—",      False, None,                  "整理番号の月部分"),
            ("serialSequence",        "整理番号_追番",    "TEXT",                False, "—",      False, None,                  "整理番号の追番部分"),
            ("barcode",               "バーコード",       "TEXT",                False, "—",      False, None,                  "整理番号を結合した文字列。UNIQUE制約あり"),
            ("vin",                   "車台番号",         "TEXT",                True,  "NULL",   False, None,                  ""),
            ("modelName",             "車種名",           "TEXT",                True,  "NULL",   False, None,                  ""),
            ("modelCode",             "型式",             "TEXT",                True,  "NULL",   False, None,                  ""),
            ("exteriorColor",         "外板色",           "TEXT",                True,  "NULL",   False, None,                  ""),
            ("inspectionType",        "点検区分",         "ENUM(InspectionType)",False, "L",      False, None,                  "L / S"),
            ("hasPolish",             "磨きあり",         "BOOLEAN",             False, "true",   False, None,                  ""),
            ("hasDressUp",            "ドレスアップあり", "BOOLEAN",             False, "false",  False, None,                  ""),
            ("hasDeliveryInspection", "完成検査あり",     "BOOLEAN",             False, "true",   False, None,                  ""),
            ("mazdaDefects",          "マツダ不具合情報", "JSONB",               True,  "NULL",   False, None,                  "マツダ側から連携される不具合内容（最大5件）"),
            ("status",                "ステータス",       "ENUM(VehicleStatus)", False, "ARRIVED",False, None,                  "ARRIVED→IN_STORAGE→IN_PROCESS→COMPLETED→DISPATCHED"),
            ("storageLocationId",     "保管場所ID",       "TEXT",                True,  "NULL",   False, "StorageLocation.id",  "UNIQUE制約あり（1対1）"),
            ("createdAt",             "作成日時",         "TIMESTAMP",           False, "now()",  False, None,                  ""),
            ("updatedAt",             "更新日時",         "TIMESTAMP",           False, "now()",  False, None,                  "更新時に自動更新"),
        ]
    },
    {
        "name": "ImportBatch", "ja": "インポートバッチ", "category": "車両・計画系",
        "desc": "作業計画データのCSVインポート履歴。誰がいつどのファイルを取り込んだかを管理する。",
        "columns": [
            ("id",           "ID",           "TEXT",              False, "cuid()", True,  None,      "CUID形式の主キー"),
            ("fileName",     "ファイル名",   "TEXT",              False, "—",      False, None,      "インポートしたCSVファイル名"),
            ("importedAt",   "インポート日時","TIMESTAMP",        False, "now()",  False, None,      ""),
            ("importedById", "インポート者ID","TEXT",             False, "—",      False, "User.id", ""),
            ("recordCount",  "件数",         "INTEGER",           False, "—",      False, None,      "インポートされた作業計画の件数"),
            ("processType",  "工程区分",     "ENUM(ProcessType)", True,  "NULL",   False, None,      "一括インポートの場合の工程"),
        ]
    },
    {
        "name": "WorkPlan", "ja": "作業計画", "category": "車両・計画系",
        "desc": "誰がいつどの車両に何の作業をするかの計画データ。CSVインポートまたは手動登録で作成される。",
        "columns": [
            ("id",               "ID",           "TEXT",              False, "cuid()", True,  None,            "CUID形式の主キー"),
            ("vehicleId",        "車両ID",       "TEXT",              False, "—",      False, "Vehicle.id",    ""),
            ("processType",      "工程区分",     "ENUM(ProcessType)", False, "—",      False, None,            "14工程のいずれか"),
            ("teamId",           "班ID",         "TEXT",              True,  "NULL",   False, "Team.id",       "担当班"),
            ("assignedWorkerId", "担当者ID",     "TEXT",              True,  "NULL",   False, "User.id",       "担当作業者"),
            ("plannedStart",     "計画開始日時", "TIMESTAMP",         False, "—",      False, None,            ""),
            ("plannedEnd",       "計画終了日時", "TIMESTAMP",         False, "—",      False, None,            ""),
            ("standardMinutes",  "標準作業時間", "INTEGER",           True,  "NULL",   False, None,            "分単位"),
            ("deliveryDeadline", "納期",         "TIMESTAMP",         True,  "NULL",   False, None,            ""),
            ("importBatchId",    "インポートID", "TEXT",              True,  "NULL",   False, "ImportBatch.id","CSVインポート時のバッチID"),
            ("createdAt",        "作成日時",     "TIMESTAMP",         False, "now()",  False, None,            ""),
        ]
    },
    # ---- 作業実績系 ----
    {
        "name": "WorkLog", "ja": "作業実績", "category": "作業実績系",
        "desc": "作業の開始・中断・再開・完了の実績記録。1回の作業（開始〜完了）で1レコード。",
        "columns": [
            ("id",          "ID",           "TEXT",                False, "cuid()", True,  None,          "CUID形式の主キー"),
            ("workPlanId",  "作業計画ID",   "TEXT",                True,  "NULL",   False, "WorkPlan.id", "計画外作業の場合はNULL"),
            ("vehicleId",   "車両ID",       "TEXT",                False, "—",      False, "Vehicle.id",  ""),
            ("workerId",    "作業者ID",     "TEXT",                False, "—",      False, "User.id",     ""),
            ("processType", "工程区分",     "ENUM(ProcessType)",   False, "—",      False, None,          ""),
            ("startedAt",   "作業開始日時", "TIMESTAMP",           False, "—",      False, None,          ""),
            ("endedAt",     "作業終了日時", "TIMESTAMP",           True,  "NULL",   False, None,          "完了するまでNULL"),
            ("status",      "ステータス",   "ENUM(WorkLogStatus)", False, "ACTIVE", False, None,          "ACTIVE / PAUSED / COMPLETED"),
            ("notes",       "メモ",         "TEXT",                True,  "NULL",   False, None,          "中断理由など"),
            ("isPlanned",   "計画作業フラグ","BOOLEAN",            False, "true",   False, None,          "false = 計画外作業"),
            ("createdAt",   "作成日時",     "TIMESTAMP",           False, "now()",  False, None,          ""),
        ]
    },
    # ---- 点検・補修系 ----
    {
        "name": "InspectionReport", "ja": "点検報告書", "category": "点検・補修系",
        "desc": "L点検・S点検の報告書ヘッダー情報。1回の点検で1レコード作成される。",
        "columns": [
            ("id",             "ID",           "TEXT",                False, "cuid()", True,  None,       "CUID形式の主キー"),
            ("vehicleId",      "車両ID",       "TEXT",                False, "—",      False, "Vehicle.id",""),
            ("workerId",       "作業者ID",     "TEXT",                False, "—",      False, "User.id",  ""),
            ("inspectionType", "点検区分",     "ENUM(InspectionType)",False, "—",      False, None,       "L / S"),
            ("startedAt",      "開始日時",     "TIMESTAMP",           False, "—",      False, None,       ""),
            ("endedAt",        "終了日時",     "TIMESTAMP",           True,  "NULL",   False, None,       "完了するまでNULL"),
            ("createdAt",      "作成日時",     "TIMESTAMP",           False, "now()",  False, None,       ""),
        ]
    },
    {
        "name": "Defect", "ja": "不具合", "category": "点検・補修系",
        "desc": "点検で発見された不具合の詳細情報。展開図上の座標・写真URLを含む。",
        "columns": [
            ("id",                 "ID",           "TEXT",                False, "cuid()", True,  None,                  "CUID形式の主キー"),
            ("inspectionReportId", "点検報告書ID", "TEXT",                False, "—",      False, "InspectionReport.id", ""),
            ("vehicleId",          "車両ID",       "TEXT",                False, "—",      False, "Vehicle.id",          ""),
            ("location",           "部位",         "TEXT",                False, "—",      False, None,                  "互換用の部位ラベル"),
            ("locationX",          "X座標",        "FLOAT",               True,  "NULL",   False, None,                  "展開図上のX座標（%）"),
            ("locationY",          "Y座標",        "FLOAT",               True,  "NULL",   False, None,                  "展開図上のY座標（%）"),
            ("defectType",         "不具合種類",   "TEXT",                False, "—",      False, None,                  "線傷/面傷/塗装/へこみ/欠け/汚れ/その他"),
            ("severity",           "程度",         "ENUM(DefectSeverity)",False, "—",      False, None,                  "A（軽微）/ B（中程度）/ C（重大）"),
            ("repairMinutes",      "補修見込み時間","INTEGER",            True,  "NULL",   False, None,                  "分単位"),
            ("description",        "説明",         "TEXT",                True,  "NULL",   False, None,                  ""),
            ("photoUrl",           "写真URL",      "TEXT",                True,  "NULL",   False, None,                  "サーバー上の写真ファイルパス"),
            ("status",             "補修状況",     "ENUM(DefectStatus)",  False, "OPEN",   False, None,                  "OPEN / IN_PROGRESS / REPAIRED"),
            ("createdAt",          "作成日時",     "TIMESTAMP",           False, "now()",  False, None,                  ""),
        ]
    },
    {
        "name": "RepairLog", "ja": "補修実績", "category": "点検・補修系",
        "desc": "不具合に対する補修作業の実績ヘッダー。どの不具合を補修したかはRepairLogDefectで管理。",
        "columns": [
            ("id",        "ID",         "TEXT",      False, "cuid()", True,  None,       "CUID形式の主キー"),
            ("vehicleId", "車両ID",     "TEXT",      False, "—",      False, "Vehicle.id",""),
            ("workerId",  "作業者ID",   "TEXT",      False, "—",      False, "User.id",  ""),
            ("startedAt", "開始日時",   "TIMESTAMP", False, "—",      False, None,       ""),
            ("endedAt",   "終了日時",   "TIMESTAMP", True,  "NULL",   False, None,       ""),
            ("notes",     "メモ",       "TEXT",      True,  "NULL",   False, None,       ""),
            ("createdAt", "作成日時",   "TIMESTAMP", False, "now()",  False, None,       ""),
        ]
    },
    {
        "name": "RepairLogDefect", "ja": "補修実績-不具合紐付け", "category": "点検・補修系",
        "desc": "補修実績と不具合の中間テーブル（多対多）。1回の補修で複数の不具合を補修できる。",
        "columns": [
            ("repairLogId", "補修実績ID", "TEXT", False, "—", True, "RepairLog.id", "複合主キー"),
            ("defectId",    "不具合ID",   "TEXT", False, "—", True, "Defect.id",    "複合主キー"),
        ]
    },
    # ---- ドレスアップ・チェック系 ----
    {
        "name": "VehicleDressUpOrder", "ja": "ドレスアップ指示", "category": "ドレスアップ・チェック系",
        "desc": "車両に対するドレスアップ作業の指示・実績。",
        "columns": [
            ("id",           "ID",         "TEXT",              False, "cuid()", True,  None,             "CUID形式の主キー"),
            ("vehicleId",    "車両ID",     "TEXT",              False, "—",      False, "Vehicle.id",     ""),
            ("dressUpItemId","ドレスアップ品ID","TEXT",          False, "—",      False, "DressUpItem.id", ""),
            ("workerId",     "作業者ID",   "TEXT",              True,  "NULL",   False, "User.id",        "担当者未定の場合はNULL"),
            ("status",       "ステータス", "ENUM(OrderStatus)", False, "PENDING",False, None,             "PENDING / IN_PROGRESS / COMPLETED"),
            ("startedAt",    "開始日時",   "TIMESTAMP",         True,  "NULL",   False, None,             ""),
            ("endedAt",      "終了日時",   "TIMESTAMP",         True,  "NULL",   False, None,             ""),
            ("createdAt",    "作成日時",   "TIMESTAMP",         False, "now()",  False, None,             ""),
        ]
    },
    {
        "name": "CheckResult", "ja": "チェック結果", "category": "ドレスアップ・チェック系",
        "desc": "各工程のチェックリスト実施結果。チェック項目と結果をJSONで保持する。",
        "columns": [
            ("id",          "ID",         "TEXT",              False, "cuid()", True,  None,         "CUID形式の主キー"),
            ("vehicleId",   "車両ID",     "TEXT",              False, "—",      False, "Vehicle.id", ""),
            ("processType", "工程区分",   "ENUM(ProcessType)", False, "—",      False, None,         ""),
            ("referenceId", "関連ID",     "TEXT",              True,  "NULL",   False, None,         "関連するWorkLog ID等"),
            ("workerId",    "作業者ID",   "TEXT",              False, "—",      False, "User.id",    ""),
            ("items",       "チェック結果","JSONB",            False, "—",      False, None,         "チェック項目と結果のJSON"),
            ("checkedAt",   "チェック日時","TIMESTAMP",        False, "now()",  False, None,         ""),
        ]
    },
    {
        "name": "Photo", "ja": "写真", "category": "ドレスアップ・チェック系",
        "desc": "チェック結果や不具合に紐付く写真ファイルの管理テーブル。",
        "columns": [
            ("id",            "ID",           "TEXT",      False, "cuid()", True,  None,             "CUID形式の主キー"),
            ("checkResultId", "チェック結果ID","TEXT",     True,  "NULL",   False, "CheckResult.id", "NULLの場合は単独写真"),
            ("workerId",      "撮影者ID",     "TEXT",      False, "—",      False, "User.id",        ""),
            ("filePath",      "ファイルパス", "TEXT",      False, "—",      False, None,             "サーバー上の保存パス"),
            ("takenAt",       "撮影日時",     "TIMESTAMP", False, "now()",  False, None,             ""),
        ]
    },
]

# ============================================================
# ENUM定義
# ============================================================
ENUMS = [
    ("Role",           "ロール",       [
        ("GENERAL",     "一般作業者", "通常の作業者"),
        ("TEAM_LEADER", "班長",       "チーム管理・進捗確認"),
        ("MANAGER",     "管理者",     "全体管理・データ管理"),
    ]),
    ("InspectionType", "点検区分",     [
        ("L", "L点検", "標準点検"),
        ("S", "S点検", "簡易点検"),
    ]),
    ("VehicleStatus",  "車両ステータス",[
        ("ARRIVED",    "入庫",           "車両が入庫した状態"),
        ("IN_STORAGE", "保管中",         "保管場所に配置済み"),
        ("IN_PROCESS", "作業中",         "何らかの作業が進行中"),
        ("COMPLETED",  "完成検査完了",   "すべての作業が完了"),
        ("DISPATCHED", "出庫済",         "車両が出庫した状態"),
    ]),
    ("ProcessType",    "工程区分",      [
        ("LOADING",          "積み込み・降ろし", ""),
        ("STORAGE",          "保管",             ""),
        ("RETRIEVAL",        "回送",             ""),
        ("WASH",             "洗車",             ""),
        ("POLISH",           "磨き",             ""),
        ("L_INSPECTION",     "L点検",            ""),
        ("S_INSPECTION",     "S点検",            ""),
        ("REPAIR",           "補修",             ""),
        ("RUST_PREVENTION",  "防錆",             ""),
        ("PARTS_ISSUE",      "ドレスアップ部品払い出し", ""),
        ("DRESS_UP",         "ドレスアップ",     ""),
        ("COATING",          "コーティング",     ""),
        ("FINAL_INSPECTION", "完成検査",         ""),
        ("OTHER",            "その他",           "計画外作業"),
    ]),
    ("WorkLogStatus",  "作業実績ステータス", [
        ("ACTIVE",    "作業中", "現在進行中"),
        ("PAUSED",    "一時中断", "中断状態"),
        ("COMPLETED", "完了",   "作業完了"),
    ]),
    ("DefectSeverity", "不具合程度",   [
        ("A", "軽微",   "軽微な不具合"),
        ("B", "中程度", "中程度の不具合"),
        ("C", "重大",   "重大な不具合"),
    ]),
    ("DefectStatus",   "補修状況",     [
        ("OPEN",        "未補修",  "未対応"),
        ("IN_PROGRESS", "補修中",  "補修作業中"),
        ("REPAIRED",    "補修済",  "補修完了"),
    ]),
    ("OrderStatus",    "ドレスアップ指示ステータス", [
        ("PENDING",     "未着手",   ""),
        ("IN_PROGRESS", "作業中",   ""),
        ("COMPLETED",   "完了",     ""),
    ]),
]

# ============================================================
# スタイルヘルパー
# ============================================================
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(name="游ゴシック", size=10, bold=False, color="000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    m = Side(style="medium", color="999999")
    s = Side(style="thin", color=C_BORDER)
    return Border(left=m, right=m, top=m, bottom=m)

def apply(ws, row, col, value="", bg=None, fnt=None, aln=None, bdr=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:  cell.fill = fill(bg)
    if fnt: cell.font = fnt
    if aln: cell.alignment = aln
    if bdr: cell.border = bdr
    return cell

def merge_apply(ws, r1, c1, r2, c2, value="", bg=None, fnt=None, aln=None, bdr=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = apply(ws, r1, c1, value, bg, fnt, aln, bdr)
    # 結合セルの残りにも背景を
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            if not (r == r1 and c == c1):
                ws.cell(r, c).fill = fill(bg or C_WHITE)
    return cell


# ============================================================
# シート生成
# ============================================================

def make_cover(wb):
    ws = wb.create_sheet("概要")
    ws.sheet_view.showGridLines = False

    # 列幅
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 3
    ws.row_dimensions[1].height = 12

    # タイトルブロック
    ws.merge_cells("B2:E4")
    apply(ws, 2, 2, "現場作業管理システム（仮称）\nテーブル定義書",
          bg=C_NAVY,
          fnt=font(size=18, bold=True, color=C_WHITE),
          aln=align("center", "center", wrap=True))
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[4].height = 18

    apply(ws, 5, 2, "作成日：2026年4月　／　情報システム部",
          bg=C_LIGHT,
          fnt=font(size=10, color=C_NAVY),
          aln=align("right", "center"))
    ws.merge_cells("B5:E5")
    ws.row_dimensions[5].height = 20

    ws.row_dimensions[6].height = 12

    # テーブル一覧ヘッダー
    headers = ["テーブル名（物理）", "テーブル名（論理）", "カテゴリ", "説明"]
    for i, h in enumerate(headers):
        apply(ws, 7, 2 + i, h,
              bg=C_NAVY,
              fnt=font(bold=True, color=C_WHITE),
              aln=align("center", "center"),
              bdr=thin_border())
    ws.row_dimensions[7].height = 22

    for row_i, tbl in enumerate(TABLES):
        r = 8 + row_i
        cat = tbl["category"]
        color = CATEGORY_COLOR.get(cat, C_GRAY_H)
        bg_row = C_GRAY_BG if row_i % 2 == 0 else C_WHITE

        apply(ws, r, 2, tbl["name"],
              bg=bg_row,
              fnt=font(bold=True, color=color),
              aln=align("left", "center"),
              bdr=thin_border())

        # ハイパーリンクで各シートへ
        cell = ws.cell(r, 2)
        cell.hyperlink = f"#{tbl['name']}!A1"
        cell.font = Font(name="游ゴシック", size=10, bold=True,
                         color=color, underline="single")

        apply(ws, r, 3, tbl["ja"],
              bg=bg_row, fnt=font(), aln=align("left", "center"), bdr=thin_border())
        apply(ws, r, 4, cat,
              bg=bg_row, fnt=font(color=color), aln=align("center", "center"), bdr=thin_border())
        apply(ws, r, 5, tbl["desc"],
              bg=bg_row, fnt=font(), aln=align("left", "center", wrap=True), bdr=thin_border())
        ws.row_dimensions[r].height = 30

    # カテゴリ凡例
    legend_row = 8 + len(TABLES) + 2
    apply(ws, legend_row, 2, "【カテゴリ凡例】",
          fnt=font(bold=True, color=C_GRAY_H))
    for i, (cat, color) in enumerate(CATEGORY_COLOR.items()):
        r = legend_row + 1 + i
        apply(ws, r, 2, cat,
              bg=color,
              fnt=font(bold=True, color=C_WHITE),
              aln=align("center", "center"),
              bdr=thin_border())
        ws.row_dimensions[r].height = 18


def make_table_sheet(wb, tbl):
    ws = wb.create_sheet(tbl["name"])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    # 列幅
    col_widths = [3, 5, 22, 20, 22, 8, 12, 20, 28, 3]
    col_labels = ["", "#", "カラム名（物理名）", "カラム名（論理名）",
                  "データ型", "NULL可", "デフォルト値", "FK参照先", "備考", ""]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    cat = tbl["category"]
    hdr_color = CATEGORY_COLOR.get(cat, C_NAVY)

    # タイトル行
    ws.merge_cells("B1:I1")
    apply(ws, 1, 2,
          f"{tbl['name']}　（{tbl['ja']}）　―　{tbl['category']}",
          bg=hdr_color,
          fnt=font(size=13, bold=True, color=C_WHITE),
          aln=align("left", "center"))
    ws.row_dimensions[1].height = 26

    # 説明行
    ws.merge_cells("B2:I2")
    apply(ws, 2, 2, tbl["desc"],
          bg=C_LIGHT,
          fnt=font(size=10, color=C_NAVY, italic=True),
          aln=align("left", "center", wrap=True))
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 6

    # カラムヘッダー
    for i, label in enumerate(col_labels):
        if label:
            apply(ws, 4, i + 1, label,
                  bg=hdr_color,
                  fnt=font(bold=True, color=C_WHITE, size=10),
                  aln=align("center", "center"),
                  bdr=thin_border())
    ws.row_dimensions[4].height = 22

    # PKカラムか判定
    pk_cols = [c[0] for c in tbl["columns"] if c[5]]

    # カラムデータ
    for row_i, col in enumerate(tbl["columns"]):
        r = 5 + row_i
        (phys, logical, dtype, nullable, default, pk, fk, notes) = col
        is_pk = pk
        is_fk = bool(fk)
        bg_row = RGBColor = C_LIGHT if is_pk else (C_GRAY_BG if row_i % 2 == 0 else C_WHITE)

        apply(ws, r, 2, row_i + 1,
              bg=bg_row, fnt=font(size=9, color="888888"), aln=align("center", "center"), bdr=thin_border())

        # 物理名（PKは太字）
        pk_mark = " 🔑" if is_pk else ("  ↗" if is_fk else "")
        apply(ws, r, 3, phys + pk_mark,
              bg=bg_row,
              fnt=font(bold=is_pk, color=hdr_color if is_pk else DARK if not is_fk else "0E6E75"),
              aln=align("left", "center"), bdr=thin_border())

        apply(ws, r, 4, logical,
              bg=bg_row, fnt=font(), aln=align("left", "center"), bdr=thin_border())
        apply(ws, r, 5, dtype,
              bg=bg_row, fnt=font(size=9, color="444444"), aln=align("left", "center"), bdr=thin_border())
        apply(ws, r, 6, "○" if nullable else "—",
              bg=bg_row,
              fnt=font(color="C0392B" if nullable else "1E8B4C", bold=True),
              aln=align("center", "center"), bdr=thin_border())
        apply(ws, r, 7, default,
              bg=bg_row, fnt=font(size=9), aln=align("center", "center"), bdr=thin_border())
        apply(ws, r, 8, fk or "",
              bg=bg_row,
              fnt=font(size=9, color="0E6E75" if fk else "000000", italic=bool(fk)),
              aln=align("left", "center"), bdr=thin_border())
        apply(ws, r, 9, notes,
              bg=bg_row, fnt=font(size=9), aln=align("left", "center", wrap=True), bdr=thin_border())
        ws.row_dimensions[r].height = 22

    # 戻るリンク
    back_row = 5 + len(tbl["columns"]) + 1
    ws.row_dimensions[back_row].height = 16
    cell = ws.cell(back_row + 1, 2, "← 概要シートに戻る")
    cell.hyperlink = "#概要!A1"
    cell.font = Font(name="游ゴシック", size=9, color=C_NAVY, underline="single")


DARK = "333333"

def make_enum_sheet(wb):
    ws = wb.create_sheet("ENUM定義")
    ws.sheet_view.showGridLines = False

    col_widths = [3, 22, 18, 22, 28, 3]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.merge_cells("B1:E1")
    apply(ws, 1, 2, "ENUM（列挙型）定義一覧",
          bg=C_NAVY,
          fnt=font(size=14, bold=True, color=C_WHITE),
          aln=align("left", "center"))
    ws.row_dimensions[1].height = 28

    current_row = 3

    for enum_name, enum_ja, values in ENUMS:
        # ENUMヘッダー
        ws.merge_cells(
            start_row=current_row, start_column=2,
            end_row=current_row, end_column=5
        )
        apply(ws, current_row, 2,
              f"{enum_name}　（{enum_ja}）",
              bg=C_NAVY_L,
              fnt=font(bold=True, color=C_WHITE, size=11),
              aln=align("left", "center"))
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # 列ヘッダー
        for ci, h in enumerate(["値（物理名）", "値（論理名）", "説明"]):
            apply(ws, current_row, 2 + ci, h,
                  bg=C_LIGHT,
                  fnt=font(bold=True, color=C_NAVY),
                  aln=align("center", "center"),
                  bdr=thin_border())
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for vi, (val, val_ja, desc) in enumerate(values):
            bg_row = C_GRAY_BG if vi % 2 == 0 else C_WHITE
            apply(ws, current_row, 2, val,
                  bg=bg_row, fnt=font(bold=True, color=DARK),
                  aln=align("left", "center"), bdr=thin_border())
            apply(ws, current_row, 3, val_ja,
                  bg=bg_row, fnt=font(),
                  aln=align("left", "center"), bdr=thin_border())
            apply(ws, current_row, 4, desc,
                  bg=bg_row, fnt=font(size=9),
                  aln=align("left", "center"), bdr=thin_border())
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1  # ENUMの間に空行


# ============================================================
# メイン
# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    make_cover(wb)
    for tbl in TABLES:
        make_table_sheet(wb, tbl)
    make_enum_sheet(wb)

    wb.save(OUT)
    print(f"保存完了: {OUT}")


if __name__ == "__main__":
    main()
